#-------------------------------------------------------------------------------

def main():
    pass

if __name__ == '__main__':
    main()

import openpyxl
from datetime import datetime

# Open the Excel file
wb = openpyxl.load_workbook('filename.xlsx')

# Select the sheet you want to modify
ws = wb['Tabelle1']

# Refresh the date in cell B1
ws['B1'].value = datetime.now()

# Placeholder replacements
replacements = {
    '[mass%]': None,
    '[FSM]': None,
    '[Polymer]': None,
    '[Temperature]': None
}

# Keep track of the user input for each unique replacement
user_input = {}

# Iterate through all the cells in the sheet
for row in ws.iter_rows():
    for cell in row:
        # Check if the cell contains a placeholder
        if type(cell.value) == str:
            for replacement_string in replacements.keys():
                if replacement_string in cell.value:
                    # Check if user input has already been obtained for this placeholder value
                    if replacement_string in user_input:
                        # Use the previous user input for this placeholder value
                        user_input_value = user_input[replacement_string]
                    else:
                        # Prompt the user for input and store it for this placeholder value
                        user_input_value = input(f"Enter replacement text for {replacement_string}: ")
                        user_input[replacement_string] = user_input_value

                    # Replace the placeholder with the user input
                    cell.value = cell.value.replace(replacement_string, user_input_value)
                    replacements[replacement_string] = user_input_value
        elif type(cell.value) == datetime:
            cell.value = cell.value.strftime('%d-%m-%Y')

# Print the replacements dictionary to check that the input values are stored correctly
print(replacements)

# Save the modified Excel file with filename
filename = f"filename_{replacements['[Polymer]']}_{replacements['[mass%]']}_{replacements['[FSM]']}.xlsx"
wb.save(filename)
